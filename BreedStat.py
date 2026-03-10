# ==================================================
# BreedStat Pro v1.0
# Complete Professional Clean Version
# ==================================================

import os
import sys
import time
import pandas as pd
import numpy as np
from sklearn.decomposition import PCA
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import streamlit as st
import pandas as pd
import numpy as np

st.title("BreedStat Pro v1.0")

uploaded_file = st.file_uploader("Upload Excel File", type=["xlsx"])

if uploaded_file is not None:

    df = pd.read_excel(uploaded_file)

    st.write("File uploaded successfully")

else:
    st.stop()
st.title("BreedStat Pro v1.0")
st.subheader("Multi Environment Trial Analysis Software")
st.write("Developed by Dr. Girish Chougule")
from scipy import stats
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from statsmodels.formula.api import mixedlm
# ==================================================
# SAFE BASE PATH (Works for .py and .exe)
# ==================================================
if getattr(sys, 'frozen', False):
    base_path = sys._MEIPASS
else:
    base_path = os.path.dirname(os.path.abspath(__file__))

logo_png = os.path.join(base_path, "BreedStat_Pro_Logo.png")
logo_ico = os.path.join(base_path, "BreedStat_Pro_Logo.ico")

# ==================================================
# MAIN WINDOW
# ==================================================




# ==================================================
# LOAD DATA
# ==================================================


if uploaded_file is not None:
    df = pd.read_excel(uploaded_file)
else:
    st.stop()

# ==================================================
# DATA PROCESSING
# ==================================================
numeric_traits = df.select_dtypes(include="number").columns.tolist()
non_numeric_traits = df.select_dtypes(exclude="number").columns.tolist()

grouped = df.groupby(["Location Name", "Genotype Name"])

mean_numeric = grouped[numeric_traits].mean().reset_index()

mode_non_numeric = (
    df.groupby(["Location Name", "Genotype Name"], as_index=False)[non_numeric_traits]
    .agg(lambda x: x.astype(str).mode().iloc[0] if not x.astype(str).mode().empty else None)
)

rep_count = grouped.size().reset_index(name="Replication Count")

mean_df = mean_numeric.merge(mode_non_numeric,
                             on=["Location Name", "Genotype Name"],
                             how="left")

mean_df = mean_df.merge(rep_count,
                        on=["Location Name", "Genotype Name"],
                        how="left")

# ==================================================
# CUSTOM CALCULATIONS
# ==================================================
mean_df["ARFCTR"] = np.where(
    (mean_df["RxR_SPA"] * mean_df["PxP_SPA"] * mean_df["PLTCNT"]) != 0,
    (100000000 / (mean_df["RxR_SPA"] * mean_df["PxP_SPA"])) / mean_df["PLTCNT"],
    np.nan
)

mean_df["GPTYLD"] = np.where(
    mean_df["GRNMST"] != 87,
    (100 - mean_df["GRNMST"]) * mean_df["GRN_WT"] / 87,
    np.nan
)

mean_df["GRNYLD"] = mean_df["ARFCTR"] * mean_df["GPTYLD"] / 100

overall_means = mean_df[["GRNYLD", "DYSSHD"]].mean()

mean_df["GRNYLD_%Mean"] = (mean_df["GRNYLD"] /
                           overall_means["GRNYLD"]) * 100

mean_df["DYSSHD_%Mean"] = (mean_df["DYSSHD"] /
                           overall_means["DYSSHD"]) * 100

mean_df["Rating"] = mean_df["GRNYLD_%Mean"] - mean_df["DYSSHD_%Mean"]

# ==================================================
# GENOTYPE AVERAGE & RANKING
# ==================================================
avg_df = mean_df.groupby("Genotype Name", as_index=False).mean(numeric_only=True)

avg_df["GRNYLD_Rank"] = avg_df["GRNYLD"].rank(method="dense",
                                             ascending=False).astype(int)

avg_df["Rating_Rank"] = avg_df["Rating"].rank(method="dense",
                                              ascending=False).astype(int)

# ==================================================
# BLUP ANALYSIS (Mixed Model)
# ==================================================
# Create safe column names for mixed model
# ==================================================
# BLUP ANALYSIS (Mixed Model - Corrected)
# ==================================================

# ==================================================
# BLUP ANALYSIS (Mixed Model - Stable Version)
# ==================================================

blup_results = []

try:

    blup_data = df.copy()

    # Calculate GRNYLD at replication level
    blup_data["ARFCTR"] = np.where(
        (blup_data["RxR_SPA"] * blup_data["PxP_SPA"] * blup_data["PLTCNT"]) != 0,
        (100000000 / (blup_data["RxR_SPA"] * blup_data["PxP_SPA"])) / blup_data["PLTCNT"],
        np.nan
    )

    blup_data["GPTYLD"] = np.where(
        blup_data["GRNMST"] != 87,
        (100 - blup_data["GRNMST"]) * blup_data["GRN_WT"] / 87,
        np.nan
    )

    blup_data["GRNYLD"] = blup_data["ARFCTR"] * blup_data["GPTYLD"] / 100

    # Remove invalid values
    blup_data = blup_data.replace([np.inf, -np.inf], np.nan)
    blup_data = blup_data.dropna(subset=["GRNYLD"])

    # Rename columns for statsmodels safety
    blup_data = blup_data.rename(columns={
        "Location Name": "Location",
        "Genotype Name": "Genotype"
    })

    # Fit Mixed Model
    model = mixedlm(
        "GRNYLD ~ C(Location)",
        data=blup_data,
        groups=blup_data["Genotype"]
    )

    result = model.fit(reml=True)

    random_effects = result.random_effects

    for genotype, value in random_effects.items():

        if isinstance(value, (list, np.ndarray)):
            blup_val = value[0]
        else:
            blup_val = float(value)

        blup_results.append([genotype, blup_val])

    blup_df = pd.DataFrame(
        blup_results,
        columns=["Genotype_Name", "BLUP_Estimate"]
    )

    blup_df = blup_df.sort_values(
        by="BLUP_Estimate",
        ascending=False
    )

    blup_df["BLUP_Rank"] = blup_df["BLUP_Estimate"].rank(
        method="dense",
        ascending=False
    ).astype(int)

except Exception as e:

    print("BLUP analysis failed:", e)

    blup_df = pd.DataFrame(
        columns=["Genotype_Name", "BLUP_Estimate", "BLUP_Rank"]
    )
# ==================================================
# STABILITY ANALYSIS
# ==================================================
# ==================================================
# STABILITY ANALYSIS (NumPy Version – No statsmodels)
# ==================================================
# ==================================================
# CREATE ENVIRONMENT INDEX
# ==================================================

env_means = (
    mean_df
    .groupby("Location Name")["GRNYLD"]
    .mean()
    .rename("Env_Index")
)

mean_df = mean_df.merge(
    env_means,
    on="Location Name",
    how="left"
)
stability_results = []

for g in mean_df["Genotype Name"].unique():

    g_data = mean_df[mean_df["Genotype Name"] == g]

    if len(g_data) < 3:
        continue

    X = g_data["Env_Index"].values
    Y = g_data["GRNYLD"].values

    # Linear regression
    slope, intercept = np.polyfit(X, Y, 1)

    Y_pred = slope * X + intercept

    # R2
    ss_total = np.sum((Y - np.mean(Y))**2)
    ss_residual = np.sum((Y - Y_pred)**2)
    r2 = 1 - (ss_residual / ss_total)

    # Residual variance
    deviation_var = np.var(Y - Y_pred)

    # ---------- P VALUE CALCULATION ----------
    n = len(X)
    x_mean = np.mean(X)

    se_slope = np.sqrt(
        ss_residual / (n - 2)
    ) / np.sqrt(np.sum((X - x_mean)**2))

    t_stat = slope / se_slope

    p_value = 2 * (1 - stats.t.cdf(abs(t_stat), df=n-2))
    # ----------------------------------------

    stability_results.append([
        g,
        slope,
        deviation_var,
        r2,
        p_value
    ])

stability_df = pd.DataFrame(
    stability_results,
    columns=[
        "Genotype Name",
        "Stability_Slope",
        "Deviation_Var",
        "R2",
        "Slope_pvalue"
    ]
)
# ==================================================
# STABILITY REMARK
# ==================================================

stability_df["Stability_Remark"] = ""

for i, row in stability_df.iterrows():

    slope = row["Stability_Slope"]
    dev = row["Deviation_Var"]

    if abs(slope - 1) <= 0.2 and dev < stability_df["Deviation_Var"].mean():
        remark = "Stable (Wide Adaptation)"

    elif slope > 1 and dev < stability_df["Deviation_Var"].mean():
        remark = "Responsive to Good Environment"

    elif slope < 1 and dev < stability_df["Deviation_Var"].mean():
        remark = "Suitable for Poor Environment"

    else:
        remark = "Unstable"

    stability_df.at[i, "Stability_Remark"] = remark
# ==================================================
# REGRESSION PLOT (Selected Test vs Multiple Checks)
# ==================================================

# ==================================================
# GENOTYPE RECOMMENDATION ANALYSIS
# ==================================================

recommend_df = pd.merge(
    avg_df,
    blup_df,
    left_on="Genotype Name",
    right_on="Genotype_Name",
    how="left"
)

recommend_df = pd.merge(
    recommend_df,
    stability_df[["Genotype Name","Stability_Remark"]],
    on="Genotype Name",
    how="left"
)

# Combined Score (lower is better)
recommend_df["Combined_Score"] = (
    recommend_df["GRNYLD_Rank"] +
    recommend_df["BLUP_Rank"]
)

recommend_df = recommend_df.sort_values("Combined_Score")

# Recommendation logic
recommendations = []

for _, row in recommend_df.iterrows():

    stability = row["Stability_Remark"]

    if stability == "Stable (Wide Adaptation)":
        rec = "Highly Recommended"

    elif stability == "Responsive to Good Environment":
        rec = "Recommended for High Input Areas"

    elif stability == "Suitable for Poor Environment":
        rec = "Recommended for Low Input Areas"

    else:
        rec = "Not Recommended"

    recommendations.append(rec)

recommend_df["Recommendation"] = recommendations

# Top 5 genotypes
top5_df = recommend_df.head(5)

# ------------------------------
# Selection Window for Regression
# ------------------------------
genotype_list = sorted(mean_df["Genotype Name"].unique())

test_entry = st.selectbox("Select Test Entry", genotype_list)

check1 = st.selectbox("Select Check Entry 1", genotype_list)
check2 = st.selectbox("Select Check Entry 2", genotype_list)
check3 = st.selectbox("Select Check Entry 3", genotype_list)

selected_checks = [c for c in [check1, check2, check3] if c]

# Combine Test + Checks
selected_genotypes = [test_entry]

if check1:
    selected_genotypes.append(check1)

if check2:
    selected_genotypes.append(check2)

if check3:
    selected_genotypes.append(check3)

# ------------------------------
# Regression Plot
# ------------------------------
plt.figure(figsize=(9,7))

colors = ["red","blue","green","orange","purple"]

for i, g in enumerate(selected_genotypes):

    g_data = mean_df[mean_df["Genotype Name"] == g]

    if len(g_data) < 3:
        continue

    X = g_data["Env_Index"].values
    Y = g_data["GRNYLD"].values

    slope, intercept = np.polyfit(X, Y, 1)

    Y_pred = slope * X + intercept

    ss_total = np.sum((Y - np.mean(Y))**2)
    ss_residual = np.sum((Y - Y_pred)**2)
    r2 = 1 - (ss_residual / ss_total)

    x_vals = np.linspace(X.min(), X.max(), 100)
    y_vals = intercept + slope * x_vals

    color = colors[i % len(colors)]

    if g == test_entry:

        plt.scatter(X, Y, color=color, s=120, marker="D")

        plt.plot(x_vals, y_vals,
                 linewidth=3,
                 color=color,
                 label=f"{g} (Test) | bi={slope:.2f} | R²={r2:.3f}")

    else:

        plt.scatter(X, Y, color=color, s=90)

        plt.plot(x_vals, y_vals,
                 linestyle="--",
                 linewidth=2,
                 color=color,
                 label=f"{g} (Check) | bi={slope:.2f} | R²={r2:.3f}")

plt.xlabel("Environmental Index")
plt.ylabel("GRNYLD")
plt.title("Regression Stability: Test vs Checks")
plt.legend(fontsize=9)
plt.grid(True)
plt.tight_layout()

regression_plot_file = "Regression_Test_vs_Check.png"

plt.savefig(regression_plot_file, dpi=300)
plt.close()
# ==================================================
# GGE BIPLOT
# ==================================================
gge_data = mean_df.pivot(index="Genotype Name",
                         columns="Location Name",
                         values="GRNYLD").fillna(0)

gge_centered = gge_data - gge_data.mean()

pca = PCA(n_components=2)
scores = pca.fit_transform(gge_centered)

plt.figure(figsize=(8, 6))
plt.scatter(scores[:, 0], scores[:, 1])

for i, genotype in enumerate(gge_data.index):
    plt.text(scores[i, 0], scores[i, 1], genotype, fontsize=8)

plt.xlabel("PC1")
plt.ylabel("PC2")
plt.title("GGE Biplot (Genotype × Environment)")
plt.grid(True)
plt.tight_layout()

gge_plot_file = "GGE_Biplot.png"
plt.savefig(gge_plot_file, dpi=300)
plt.close()



# ==================================================
# BLUP + STABILITY GRAPH SHEET
# ==================================================


# ==================================================
# BLUP + STABILITY GRAPH
# ==================================================

# ==================================================
# BLUP + STABILITY GRAPH
# ==================================================

blup_stability_plot = "BLUP_Stability_Graph.png"

if not blup_df.empty and not stability_df.empty:

    merged_df = pd.merge(
        blup_df,
        stability_df,
        left_on="Genotype_Name",
        right_on="Genotype Name",
        how="inner"
    )

    plt.figure(figsize=(9,7))

    x = merged_df["Stability_Slope"]
    y = merged_df["BLUP_Estimate"]

    plt.scatter(x, y, s=80)

    for i, row in merged_df.iterrows():
        plt.text(
            row["Stability_Slope"],
            row["BLUP_Estimate"],
            row["Genotype_Name"],
            fontsize=8
        )

    plt.axvline(1, linestyle="--")
    plt.axhline(0, linestyle="--")

    plt.xlabel("Stability Slope (bi)")
    plt.ylabel("BLUP Estimate")
    plt.title("BLUP vs Stability Relationship")

    plt.grid(True)
    plt.tight_layout()

    plt.savefig(blup_stability_plot, dpi=300)
    plt.close()


# ==================================================
# AMMI ANALYSIS
# ==================================================

ammi_matrix = mean_df.pivot(
    index="Genotype Name",
    columns="Location Name",
    values="GRNYLD"
)

grand_mean = ammi_matrix.values.mean()

env_means = ammi_matrix.mean(axis=0)
geno_means = ammi_matrix.mean(axis=1)

interaction = ammi_matrix.values - geno_means.values[:, None] - env_means.values + grand_mean

interaction = pd.DataFrame(
    interaction,
    index=ammi_matrix.index,
    columns=ammi_matrix.columns
)

interaction = interaction.fillna(0)

# PCA for IPCA
pca_ammi = PCA(n_components=2)
ipca_scores = pca_ammi.fit_transform(interaction)

ipca1 = ipca_scores[:, 0]
ipca2 = ipca_scores[:, 1]

ammi_df = pd.DataFrame({
    "Genotype Name": ammi_matrix.index,
    "Mean_Yield": geno_means.values,
    "IPCA1": ipca1,
    "IPCA2": ipca2
})

# ---------------- AMMI1 BIPLOT ----------------
plt.figure(figsize=(8,6))

plt.scatter(ammi_df["Mean_Yield"], ammi_df["IPCA1"], s=80)

for i,row in ammi_df.iterrows():
    plt.text(row["Mean_Yield"], row["IPCA1"], row["Genotype Name"], fontsize=8)

plt.axhline(0, linestyle="--")

plt.xlabel("Mean Yield")
plt.ylabel("IPCA1")
plt.title("AMMI1 Biplot")
plt.grid(True)
plt.tight_layout()

ammi1_plot = "AMMI1_Biplot.png"
plt.savefig(ammi1_plot, dpi=300)
plt.close()

# ---------------- AMMI2 BIPLOT ----------------
plt.figure(figsize=(8,6))

plt.scatter(ammi_df["IPCA1"], ammi_df["IPCA2"], s=80)

for i,row in ammi_df.iterrows():
    plt.text(row["IPCA1"], row["IPCA2"], row["Genotype Name"], fontsize=8)

plt.axhline(0, linestyle="--")
plt.axvline(0, linestyle="--")

plt.xlabel("IPCA1")
plt.ylabel("IPCA2")
plt.title("AMMI2 Biplot")
plt.grid(True)
plt.tight_layout()

ammi2_plot = "AMMI2_Biplot.png"
plt.savefig(ammi2_plot, dpi=300)
plt.close()

# ==================================================
# AMMI STABILITY VALUE (ASV)
# ==================================================

ipca1_var = pca_ammi.explained_variance_ratio_[0]
ipca2_var = pca_ammi.explained_variance_ratio_[1]

asv = np.sqrt(
    (ammi_df["IPCA1"] / ipca1_var)**2 +
    (ammi_df["IPCA2"] / ipca2_var)**2
)

ammi_df["ASV"] = asv
ammi_df["ASV_Rank"] = ammi_df["ASV"].rank(method="dense").astype(int)

# ==================================================
# GGE WHICH WON WHERE POLYGON
# ==================================================

from scipy.spatial import ConvexHull

plt.figure(figsize=(8,6))

x = scores[:,0]
y = scores[:,1]

plt.scatter(x,y)

for i, g in enumerate(gge_data.index):
    plt.text(x[i], y[i], g, fontsize=8)

points = np.column_stack((x,y))

if len(points) >= 3:

    hull = ConvexHull(points)

    for simplex in hull.simplices:
        plt.plot(points[simplex,0], points[simplex,1], 'r-')

plt.xlabel("PC1")
plt.ylabel("PC2")
plt.title("GGE Which Won Where")
plt.grid(True)
plt.tight_layout()

gge_polygon_plot = "GGE_WhichWonWhere.png"

plt.savefig(gge_polygon_plot, dpi=300)
plt.close()


# -------------------------
# AMMI Stability Ranking
# -------------------------
# ==================================================
# SAVE OUTPUT
# ==================================================

output_file = "BreedStat_Pro_Output.xlsx"
with open(output_file, "rb") as f:
    st.download_button(
        "Download Excel Report",
        f,
        file_name="BreedStat_Pro_Output.xlsx"
    )

with pd.ExcelWriter(output_file, engine="openpyxl") as writer:

    mean_df.to_excel(writer, sheet_name="Detailed Results", index=False)
    avg_df.to_excel(writer, sheet_name="Summary", index=False)
    stability_df.to_excel(writer, sheet_name="Stability", index=False)
    blup_df.to_excel(writer, sheet_name="BLUP_Analysis", index=False)
    ammi_df.to_excel(writer, sheet_name="AMMI_Stability", index=False)


# Load workbook to insert graphs
wb = load_workbook(output_file)

# -------------------------
# BLUP Stability Graph
# -------------------------
if os.path.exists(blup_stability_plot):

    ws_blup = wb.create_sheet("BLUP_Stability")

    img_blup = XLImage(blup_stability_plot)
    img_blup.width = 700
    img_blup.height = 500
    img_blup.anchor = "A1"

    ws_blup.add_image(img_blup)


top5_df.to_excel(writer,
                 sheet_name="Top_5_Genotypes",
                 index=False)

recommend_df.to_excel(writer,
                      sheet_name="Genotype_Recommendation",
                      index=False)

# -------------------------
# Regression Plot
# -------------------------
if os.path.exists(regression_plot_file):

    ws_reg = wb.create_sheet("Regression_Test_vs_Check")

    img_reg = XLImage(regression_plot_file)
    img_reg.width = 700
    img_reg.height = 500
    img_reg.anchor = "A1"

    ws_reg.add_image(img_reg)


# -------------------------
# GGE Biplot
# -------------------------
if os.path.exists(gge_plot_file):

    ws_gge = wb.create_sheet("GGE_Biplot")

    img_gge = XLImage(gge_plot_file)
    img_gge.width = 600
    img_gge.height = 400
    img_gge.anchor = "A1"

    ws_gge.add_image(img_gge)


# -------------------------
# AMMI1 Biplot
# -------------------------
if os.path.exists(ammi1_plot):

    ws_ammi1 = wb.create_sheet("AMMI1_Biplot")

    img_ammi1 = XLImage(ammi1_plot)
    img_ammi1.width = 600
    img_ammi1.height = 400
    img_ammi1.anchor = "A1"

    ws_ammi1.add_image(img_ammi1)


# -------------------------
# AMMI2 Biplot
# -------------------------
if os.path.exists(ammi2_plot):

    ws_ammi2 = wb.create_sheet("AMMI2_Biplot")

    img_ammi2 = XLImage(ammi2_plot)
    img_ammi2.width = 600
    img_ammi2.height = 400
    img_ammi2.anchor = "A1"

    ws_ammi2.add_image(img_ammi2)


# -------------------------
# GGE Which Won Where
# -------------------------
if os.path.exists(gge_polygon_plot):

    ws_poly = wb.create_sheet("GGE_WhichWonWhere")

    img_poly = XLImage(gge_polygon_plot)
    img_poly.width = 600
    img_poly.height = 400
    img_poly.anchor = "A1"

    ws_poly.add_image(img_poly)


wb.save(output_file)

# ==========================================
# UI MESSAGE – TOP GENOTYPE IDENTIFIED
# ==========================================

try:
    top_genotype = top5_df.iloc[0]["Genotype Name"]

    st.success("✅ BreedStat Pro Analysis Completed")

    st.info(f"""
Top Performing Genotype: **{top_genotype}**

Results saved to:
{output_file}
""")

except:
    st.success("✅ BreedStat Pro Analysis Completed")

    st.info(f"""
Results saved to:
{output_file}
""")

st.write(f"Enhanced analysis saved to: {output_file}")



